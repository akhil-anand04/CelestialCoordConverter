import openpyxl
import astropy.units as u
import astropy.coordinates as coord

def extract_data_from_excel(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active worksheet
    worksheet = workbook.active

    # Dictionary to store the extracted data
    table_data = {}

    # Iterate through rows starting from the second row (assuming the first row contains headers)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Extract data from specific columns
        target_name = row[0]
        ra_str = row[1]
        dec_str = row[2]

        # Skip rows with non-numeric or improperly formatted values
        if ra_str is None or dec_str is None:
            continue

        try:
            # Convert RA and Dec from string to decimal degrees
            ra_decimal = convert_ra_to_decimal(ra_str)
            dec_decimal = convert_dec_to_decimal(dec_str)

            # Store the values in the dictionary
            table_data[target_name] = {'R.A. (J2000)': ra_decimal, 'Decl.': dec_decimal}
        except (ValueError, IndexError):
            # Skip rows with non-numeric or improperly formatted values
            continue

    # Close the workbook when done
    workbook.close()

    return table_data


def convert_ra_to_decimal(ra_str):
    # Split RA string into components
    ra_components = ra_str.split()

    # Convert RA components to decimal degrees
    ra_decimal = (float(ra_components[0]) + float(ra_components[1])/60 + float(ra_components[2])/3600) * 15  # hours to degrees
    return ra_decimal


def convert_dec_to_decimal(dec_str):
    # Split Dec string into components
    dec_components = dec_str.split()

    # Convert Dec components to decimal degrees
    sign = 1 if dec_components[0][0] == '+' else -1
    dec_decimal = sign * ((float(dec_components[0])) + float(dec_components[1])/60 + float(dec_components[2])/3600)  # degrees
    return dec_decimal


def compute_table(table_data):
    # Extract coordinates from the result table
    ra = [entry['R.A. (J2000)'] for entry in table_data.values()]
    dec = [entry['Decl.'] for entry in table_data.values()]
    target_names = list(table_data.keys())  # Extract target names as well
    return ra, dec, target_names


def galactic(ra, dec, target_names):
    galactic_coords = []
    for i in range(len(ra)):
        # Convert Equatorial coordinates to Galactic coordinates
        c = coord.SkyCoord(ra[i], dec[i], unit=(u.deg, u.deg))
        galactic_coord = c.transform_to(coord.Galactic())
        galactic_coords.append((target_names[i], galactic_coord))  # Include target_name with Galactic coordinates

        print(f"Target {target_names[i]}: Galactic Coordinates (l, b) = {galactic_coord.l.deg}, {galactic_coord.b.deg}")

    return galactic_coords


def filter_galactic(galactic_coords):
    filtered_coords = []
    for target_name, coord in galactic_coords:
        l = coord.l.deg
        if 0 < l < 180:
            filtered_coords.append((target_name, coord))
            print(f"Target {target_name}: Galactic Coordinates (l, b) = {coord.l.deg}, {coord.b.deg}")

    return filtered_coords


def export_to_excel(filtered_coords):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a new worksheet
    worksheet = workbook.active
    worksheet.title = "Filtered Galactic Coordinates"

    # Write headers
    worksheet['A1'] = "Target Name"
    worksheet['B1'] = "Galactic Longitude (l)"
    worksheet['C1'] = "Galactic Latitude (b)"

    # Write data
    for i, (target_name, coord) in enumerate(filtered_coords, start=2):
        worksheet[f'A{i}'] = target_name
        worksheet[f'B{i}'] = coord.l.deg
        worksheet[f'C{i}'] = coord.b.deg

    # Save the workbook
    workbook.save(filename="Filtered_Galactic_Coordinates3.xlsx")


def main():
    file_path = 'Coords.xlsx'
    table_data = extract_data_from_excel(file_path)
    ra, dec, target_names = compute_table(table_data)
    galactic_coords = galactic(ra, dec, target_names)
    filtered_coords = filter_galactic(galactic_coords)
    if filtered_coords:
        export_to_excel(filtered_coords)
    else:
        print("No filtered coordinates to export.")

main()
